#!/usr/bin/env python3
"""
Compare manual vs DB totals by date + employee + customer and generate Excel report.
Manual sources: weekly exports (Week 2/3) + optional voice XLSX entries.
DB source: prod app.db filtered to weeks <= week_limit (Wednesday week start).
"""

import argparse
import csv
import sqlite3
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


DAY_NAMES = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--exports-root", required=True, help="Path to exports root")
    parser.add_argument("--weeks", default="", help="Comma-separated week folders (Week 2,Week 3)")
    parser.add_argument("--manual-xlsx", default="", help="Comma-separated XLSX files with Date/Job/Total")
    parser.add_argument("--db", required=True, help="Path to prod app.db")
    parser.add_argument("--week-limit", default="", help="Week start cutoff YYYY-MM-DD (inclusive)")
    parser.add_argument("--out", required=True, help="Output XLSX file")
    return parser.parse_args()


def normalize_text(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def normalize_employee(name: str) -> str:
    return normalize_text(name)


def normalize_customer(name: str) -> str:
    return normalize_text(name)


def round_hours(value) -> float:
    try:
        return round(float(value), 2)
    except Exception:
        return 0.0


def parse_employee_from_filename(path: Path) -> str:
    name = path.stem
    name = name.split("(")[0].strip()
    return name


def read_xls_rows(path: Path) -> List[List]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def read_xlsx_rows(path: Path) -> List[List]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def find_header_indices(rows: List[List]) -> Optional[Tuple[int, int, int, int]]:
    for idx, row in enumerate(rows[:10]):
        lowered = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "date" in lowered and "client name" in lowered and "hours per job" in lowered:
            return (idx, lowered.index("date"), lowered.index("client name"), lowered.index("hours per job"))
    return None


def is_day_name(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in DAY_NAMES


def is_day_number(value) -> bool:
    return isinstance(value, (int, float)) and 1 <= int(value) <= 31


def build_date(year: int, month: int, day: int) -> str:
    return datetime(year, month, day).strftime("%Y-%m-%d")


def parse_manual_entries(path: Path) -> List[Tuple[str, str, str, float]]:
    rows = read_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else read_xls_rows(path)
    header = find_header_indices(rows)
    if not header:
        return []
    header_row, date_col, client_col, hours_col = header
    file_mtime = datetime.utcfromtimestamp(path.stat().st_mtime)
    year = file_mtime.year
    month = file_mtime.month

    entries = []
    pending = []
    current_date = None
    employee = parse_employee_from_filename(path)

    for row in rows[header_row + 1 :]:
        if not row or len(row) <= max(date_col, client_col, hours_col):
            continue
        date_cell = row[date_col]
        client_cell = row[client_col]
        hours_cell = row[hours_col]

        client = str(client_cell).strip() if client_cell is not None else ""
        hours = round_hours(hours_cell)

        if is_day_name(date_cell):
            if client and hours:
                pending.append((employee, client, hours))
            continue

        if is_day_number(date_cell):
            try:
                day_num = int(date_cell)
                use_year, use_month = year, month
                if day_num > file_mtime.day:
                    if use_month == 1:
                        use_year -= 1
                        use_month = 12
                    else:
                        use_month -= 1
                current_date = build_date(use_year, use_month, day_num)
            except ValueError:
                current_date = None
                pending = []
                continue
            if pending:
                for emp, cust, hrs in pending:
                    entries.append((current_date, emp, cust, hrs))
                pending = []
            if client and hours:
                entries.append((current_date, employee, client, hours))
            continue

        if current_date and client and hours:
            entries.append((current_date, employee, client, hours))

    return entries


def collect_manual_entries(exports_root: Path, week_dirs: Iterable[Path]) -> List[Tuple[str, str, str, float]]:
    entries: List[Tuple[str, str, str, float]] = []
    for week_dir in week_dirs:
        for path in sorted(week_dir.glob("*.xls*")):
            entries.extend(parse_manual_entries(path))
    return entries


def parse_manual_xlsx(path: Path) -> List[Tuple[str, str, str, float]]:
    wb = load_workbook(path, data_only=True)
    entries: List[Tuple[str, str, str, float]] = []
    for ws in wb.worksheets:
        employee = ws.title.replace("_", " ").strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        if "date" not in header or "job" not in header or "total" not in header:
            continue
        date_idx = header.index("date")
        job_idx = header.index("job")
        total_idx = header.index("total")
        for row in rows[1:]:
            if not row or len(row) <= max(date_idx, job_idx, total_idx):
                continue
            date_val = row[date_idx]
            total_val = row[total_idx]
            if not date_val or total_val in (None, ""):
                continue
            date_str = str(date_val).strip()
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                date = date_str
            else:
                try:
                    date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                except Exception:
                    continue
            customer = str(row[job_idx] or "").strip() or "Unknown"
            hours = round_hours(total_val)
            if hours <= 0:
                continue
            entries.append((date, employee, customer, hours))
    return entries


def load_db_entries(db_path: Path) -> List[Tuple[str, str, str, float]]:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT te.work_date, e.name, c.name, te.hours
        FROM time_entries te
        JOIN employees e ON e.id = te.employee_id
        JOIN customers c ON c.id = te.customer_id
        """
    )
    rows = cur.fetchall()
    conn.close()
    return [(r[0], r[1], r[2], float(r[3])) for r in rows]


def aggregate_daily(entries: List[Tuple[str, str, str, float]]) -> Counter:
    counter = Counter()
    for date, emp, cust, hours in entries:
        key = (date, normalize_employee(emp), normalize_customer(cust))
        counter[key] += round_hours(hours)
    rounded = Counter()
    for key, value in counter.items():
        rounded[key] = round_hours(value)
    return rounded


def week_start(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    weekday = dt.weekday()  # Monday=0
    # Wednesday start
    delta = (weekday - 2) % 7
    start = dt - timedelta(days=delta)
    return start.strftime("%Y-%m-%d")


def filter_by_week(counter: Counter, week_limit: Optional[str]) -> Counter:
    if not week_limit:
        return counter
    limit_dt = datetime.strptime(week_limit, "%Y-%m-%d")
    filtered = Counter()
    for (date, emp, cust), hours in counter.items():
        if datetime.strptime(week_start(date), "%Y-%m-%d") <= limit_dt:
            filtered[(date, emp, cust)] = hours
    return filtered


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 2))


def main():
    args = parse_args()
    exports_root = Path(args.exports_root)
    week_dirs = []
    if args.weeks:
        for name in [w.strip() for w in args.weeks.split(",") if w.strip()]:
            week_dirs.append(exports_root / name)
    else:
        week_dirs = sorted([p for p in exports_root.iterdir() if p.is_dir() and p.name.lower().startswith("week")])

    manual_entries = collect_manual_entries(exports_root, week_dirs)
    if args.manual_xlsx:
        for path in [p.strip() for p in args.manual_xlsx.split(",") if p.strip()]:
            manual_entries.extend(parse_manual_xlsx(Path(path)))

    db_entries = load_db_entries(Path(args.db))

    manual_daily = filter_by_week(aggregate_daily(manual_entries), args.week_limit)
    db_daily = filter_by_week(aggregate_daily(db_entries), args.week_limit)

    all_keys = set(manual_daily) | set(db_daily)

    wb = Workbook()
    wb.remove(wb.active)

    red = PatternFill("solid", fgColor="FFD6D6")
    green = PatternFill("solid", fgColor="D6F5D6")

    compare = wb.create_sheet("Compare_Daily")
    compare.append(["date", "employee", "customer", "manual_hours", "db_hours", "diff", "status"])
    for (date, emp, cust) in sorted(all_keys):
        m = manual_daily.get((date, emp, cust), 0)
        d = db_daily.get((date, emp, cust), 0)
        diff = round_hours(m - d)
        status = "match" if abs(diff) < 0.01 else "needs_review"
        compare.append([date, emp, cust, m, d, diff, status])
        if status == "match":
            for cell in compare[compare.max_row]:
                cell.fill = green
        else:
            for cell in compare[compare.max_row]:
                cell.fill = red
    for cell in compare[1]:
        cell.font = Font(bold=True)
    compare.freeze_panes = "A2"
    autosize(compare)

    manual_sheet = wb.create_sheet("Manual_Daily")
    manual_sheet.append(["date", "employee", "customer", "hours"])
    for (date, emp, cust), hours in sorted(manual_daily.items()):
        manual_sheet.append([date, emp, cust, hours])
    for cell in manual_sheet[1]:
        cell.font = Font(bold=True)
    manual_sheet.freeze_panes = "A2"
    autosize(manual_sheet)

    db_sheet = wb.create_sheet("DB_Daily")
    db_sheet.append(["date", "employee", "customer", "hours"])
    for (date, emp, cust), hours in sorted(db_daily.items()):
        db_sheet.append([date, emp, cust, hours])
    for cell in db_sheet[1]:
        cell.font = Font(bold=True)
    db_sheet.freeze_panes = "A2"
    autosize(db_sheet)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
