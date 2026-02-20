#!/usr/bin/env python3
"""
Line-by-line month-to-date reconciliation:
- Parse weekly manual exports (XLS/XLSX).
- Compare against SQLite time_entries (joined to employees/customers).
Outputs CSVs for matches, manual-only, and db-only entries.
"""

import argparse
import csv
import re
import sqlite3
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import xlrd
from openpyxl import load_workbook


DAY_NAMES = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--exports-root", required=True, help="Path to jcw_payroll exports root")
    parser.add_argument("--weeks", default="", help="Comma-separated week folders (e.g. 'Week 2,Week 3')")
    parser.add_argument("--db", required=True, help="Path to payroll SQLite db (app.db)")
    parser.add_argument("--month", default="", help="Month filter (YYYY-MM). Defaults to file mtime month.")
    parser.add_argument("--out-dir", required=True, help="Output directory for CSV reports")
    parser.add_argument(
        "--manual-xlsx",
        default="",
        help="Optional comma-separated XLSX files with columns Date, Job, Task, Start, Lunch, End, Total.",
    )
    return parser.parse_args()


def normalize_text(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(value or "").strip().lower())
    return cleaned


def normalize_employee(name: str) -> str:
    return normalize_text(name)


def normalize_customer(name: str) -> str:
    return normalize_text(name)


def round_hours(value) -> float:
    try:
        return round(float(value), 2)
    except Exception:
        return 0.0


def parse_employee_from_filename(path: Path) -> str:
    name = path.stem
    name = re.sub(r"\s*\(\d+\)\s*$", "", name).strip()
    return name


def read_xls_rows(path: Path) -> List[List]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows


def read_xlsx_rows(path: Path) -> List[List]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def find_header_indices(rows: List[List]) -> Optional[Tuple[int, int, int, int]]:
    # Returns (row_index, date_col, client_col, hours_col)
    for idx, row in enumerate(rows[:10]):
        if not row:
            continue
        lowered = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "date" in lowered and "client name" in lowered and "hours per job" in lowered:
            return (idx, lowered.index("date"), lowered.index("client name"), lowered.index("hours per job"))
    return None


def is_day_name(value) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().lower() in DAY_NAMES


def is_day_number(value) -> bool:
    if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
        return True
    return False


def build_date(year: int, month: int, day: int) -> str:
    return datetime(year, month, day).strftime("%Y-%m-%d")


def parse_manual_entries(path: Path, month_hint: Optional[str] = None) -> List[Tuple[str, str, str, float]]:
    rows = read_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else read_xls_rows(path)
    header = find_header_indices(rows)
    if not header:
        return []
    header_row, date_col, client_col, hours_col = header
    file_mtime = datetime.utcfromtimestamp(path.stat().st_mtime)
    year = file_mtime.year
    month = file_mtime.month
    if month_hint:
        try:
            year, month = map(int, month_hint.split("-"))
        except Exception:
            pass

    entries = []
    pending = []
    current_date = None
    employee = parse_employee_from_filename(path)

    for row in rows[header_row + 1 :]:
        if not row or len(row) <= max(date_col, client_col, hours_col):
            continue
        date_cell = row[date_col]
        client_cell = row[client_col]
        hours_cell = row[hours_col]

        client = str(client_cell).strip() if client_cell is not None else ""
        hours = round_hours(hours_cell)

        if is_day_name(date_cell):
            if client and hours:
                pending.append((employee, client, hours))
            continue

        if is_day_number(date_cell):
            try:
                day_num = int(date_cell)
                use_year, use_month = year, month
                if not month_hint and day_num > file_mtime.day:
                    # Week spans previous month (e.g., 31 when file is Feb 9)
                    if use_month == 1:
                        use_year -= 1
                        use_month = 12
                    else:
                        use_month -= 1
                current_date = build_date(use_year, use_month, day_num)
            except ValueError:
                # Skip invalid day numbers (e.g., 30 in February)
                current_date = None
                pending = []
                continue
            if pending:
                for emp, cust, hrs in pending:
                    entries.append((current_date, emp, cust, hrs))
                pending = []
            if client and hours:
                entries.append((current_date, employee, client, hours))
            continue

        if current_date and client and hours:
            entries.append((current_date, employee, client, hours))

    # If any pending without date, skip (unknown date)
    return entries


def collect_manual_entries(exports_root: Path, week_dirs: Iterable[Path], month_hint: Optional[str]) -> List[Tuple[str, str, str, float]]:
    entries: List[Tuple[str, str, str, float]] = []
    for week_dir in week_dirs:
        for path in sorted(week_dir.glob("*.xls*")):
            entries.extend(parse_manual_entries(path, month_hint))
    return entries


def parse_manual_xlsx(path: Path) -> List[Tuple[str, str, str, float]]:
    wb = load_workbook(path, data_only=True)
    entries: List[Tuple[str, str, str, float]] = []
    for ws in wb.worksheets:
        employee = ws.title.replace("_", " ").strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            date_idx = header.index("date")
            job_idx = header.index("job")
            total_idx = header.index("total")
        except ValueError:
            continue
        for row in rows[1:]:
            if not row or len(row) <= max(date_idx, job_idx, total_idx):
                continue
            date_val = row[date_idx]
            job_val = row[job_idx]
            total_val = row[total_idx]
            if not date_val or not total_val:
                continue
            date_str = str(date_val).strip()
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                date = date_str
            else:
                try:
                    date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                except Exception:
                    continue
            customer = str(job_val or "").strip() or "Unknown"
            hours = round_hours(total_val)
            if hours <= 0:
                continue
            entries.append((date, employee, customer, hours))
    return entries


def load_db_entries(db_path: Path, month: str) -> List[Tuple[str, str, str, float]]:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT te.work_date, e.name, c.name, te.hours
        FROM time_entries te
        JOIN employees e ON e.id = te.employee_id
        JOIN customers c ON c.id = te.customer_id
        WHERE te.work_date LIKE ?
        """,
        (f"{month}-%",),
    )
    rows = cur.fetchall()
    conn.close()
    return [(r[0], r[1], r[2], float(r[3])) for r in rows]


def to_counter(entries: List[Tuple[str, str, str, float]]) -> Counter:
    counts = Counter()
    for date, emp, cust, hours in entries:
        key = (date, normalize_employee(emp), normalize_customer(cust), round_hours(hours))
        counts[key] += 1
    return counts


def write_csv(path: Path, rows: List[Tuple[str, str, str, float, int]]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["date", "employee", "customer", "hours", "count"])
        writer.writerows(rows)


def expand_counter(counter: Counter) -> List[Tuple[str, str, str, float, int]]:
    rows = []
    for (date, emp, cust, hours), count in sorted(counter.items()):
        rows.append((date, emp, cust, hours, count))
    return rows


def main():
    args = parse_args()
    exports_root = Path(args.exports_root)
    db_path = Path(args.db)
    out_dir = Path(args.out_dir)

    if not exports_root.exists():
        raise SystemExit(f"Exports root not found: {exports_root}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    week_dirs = []
    if args.weeks:
        for name in [w.strip() for w in args.weeks.split(",") if w.strip()]:
            week_dirs.append(exports_root / name)
    else:
        week_dirs = sorted([p for p in exports_root.iterdir() if p.is_dir() and p.name.lower().startswith("week")])

    for path in week_dirs:
        if not path.exists():
            raise SystemExit(f"Week folder not found: {path}")

    month_hint = args.month or ""
    manual_entries = collect_manual_entries(exports_root, week_dirs, month_hint or None)
    manual_from_voice = []
    if args.manual_xlsx:
        for path in [p.strip() for p in args.manual_xlsx.split(",") if p.strip()]:
            manual_from_voice.extend(parse_manual_xlsx(Path(path)))
    if manual_from_voice:
        manual_entries.extend(manual_from_voice)
    if not manual_entries:
        raise SystemExit("No manual entries parsed from weekly exports.")

    # Determine month from first entry if not specified
    if not month_hint:
        first_date = manual_entries[0][0]
        month_hint = first_date[:7]

    db_entries = load_db_entries(db_path, month_hint)

    manual_counter = to_counter(manual_entries)
    db_counter = to_counter(db_entries)

    common = manual_counter & db_counter
    manual_only = manual_counter - db_counter
    db_only = db_counter - manual_counter

    write_csv(out_dir / f"mtd_matches_{month_hint}.csv", expand_counter(common))
    write_csv(out_dir / f"mtd_manual_only_{month_hint}.csv", expand_counter(manual_only))
    write_csv(out_dir / f"mtd_db_only_{month_hint}.csv", expand_counter(db_only))

    summary = [
        f"Month: {month_hint}",
        f"Manual entries: {sum(manual_counter.values())}",
        f"Manual entries (voice): {len(manual_from_voice)}",
        f"DB entries: {sum(db_counter.values())}",
        f"Matches: {sum(common.values())}",
        f"Manual-only: {sum(manual_only.values())}",
        f"DB-only: {sum(db_only.values())}",
    ]
    (out_dir / f"mtd_summary_{month_hint}.txt").write_text("\n".join(summary), encoding="utf-8")
    print("\n".join(summary))
    print(f"Reports: {out_dir}")


if __name__ == "__main__":
    main()
