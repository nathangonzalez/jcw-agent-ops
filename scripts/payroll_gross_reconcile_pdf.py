#!/usr/bin/env python3
"""
Reconcile gross wages from payroll PDF reports vs DB (week-based).
"""

import argparse
import re
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATE_IN_NAME = re.compile(r"(\d{6})")
LINE_GROSS = re.compile(r"^\S+\s+([0-9]+\.?[0-9]*)\s+")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True, help="Path to prod app.db")
    parser.add_argument("--pdfs", required=True, help="Comma-separated payroll PDF paths")
    parser.add_argument("--out", required=True, help="Output XLSX report")
    return parser.parse_args()


def normalize_name(name: str) -> str:
    return " ".join(name.strip().lower().split())


def parse_employee_name(raw: str) -> str:
    raw = raw.strip()
    if "," in raw:
        last, first = [p.strip() for p in raw.split(",", 1)]
        return f"{first} {last}".strip()
    return raw


def extract_pdf_gross(path: Path) -> Dict[str, float]:
    doc = fitz.open(path)
    gross_by_emp: Dict[str, float] = defaultdict(float)
    pending_gross = None
    for page in doc:
        text = page.get_text("text")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            match = LINE_GROSS.match(line)
            if match:
                try:
                    pending_gross = float(match.group(1))
                except Exception:
                    pending_gross = None
                continue
            if pending_gross is not None and "," in line:
                emp = parse_employee_name(line)
                gross_by_emp[normalize_name(emp)] += pending_gross
                pending_gross = None
    return gross_by_emp


def load_db_rates(conn) -> Dict[str, float]:
    cur = conn.cursor()
    cur.execute("SELECT name, default_pay_rate FROM employees")
    return {normalize_name(r[0]): float(r[1] or 0) for r in cur.fetchall()}


def week_range_from_filename(path: Path) -> Tuple[str, str]:
    match = DATE_IN_NAME.search(path.name)
    if not match:
        raise ValueError(f"Missing date in filename: {path.name}")
    mmddyy = match.group(1)
    week_end = datetime.strptime(mmddyy, "%m%d%y")
    week_start = week_end - timedelta(days=6)
    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d")


def db_gross_for_week(conn, week_start: str, week_end: str, rates: Dict[str, float]) -> Dict[str, float]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT te.work_date, e.name, te.hours
        FROM time_entries te
        JOIN employees e ON e.id = te.employee_id
        WHERE te.work_date >= ? AND te.work_date <= ?
        """,
        (week_start, week_end),
    )
    gross_by_emp: Dict[str, float] = defaultdict(float)
    for _, name, hours in cur.fetchall():
        emp_key = normalize_name(name)
        rate = rates.get(emp_key, 0)
        gross_by_emp[emp_key] += float(hours or 0) * float(rate or 0)
    return gross_by_emp


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))


def main():
    args = parse_args()
    db_path = Path(args.db)
    pdf_paths = [Path(p.strip()) for p in args.pdfs.split(",") if p.strip()]
    out_path = Path(args.out)

    conn = sqlite3.connect(db_path)
    rates = load_db_rates(conn)

    wb = Workbook()
    wb.remove(wb.active)
    green = PatternFill("solid", fgColor="D6F5D6")
    red = PatternFill("solid", fgColor="FFD6D6")

    summary = wb.create_sheet("Summary")
    summary.append(["pdf", "week_start", "week_end", "pdf_gross_total", "db_gross_total", "diff"])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    for pdf in pdf_paths:
        week_start, week_end = week_range_from_filename(pdf)
        pdf_gross = extract_pdf_gross(pdf)
        db_gross = db_gross_for_week(conn, week_start, week_end, rates)

        sheet_name = pdf.stem[:28]
        ws = wb.create_sheet(sheet_name)
        ws.append(["employee", "pdf_gross", "db_gross", "diff"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        all_emps = set(pdf_gross) | set(db_gross)
        pdf_total = 0.0
        db_total = 0.0
        for emp in sorted(all_emps):
            pdf_val = round(float(pdf_gross.get(emp, 0)), 2)
            db_val = round(float(db_gross.get(emp, 0)), 2)
            diff = round(pdf_val - db_val, 2)
            pdf_total += pdf_val
            db_total += db_val
            ws.append([emp, pdf_val, db_val, diff])
            row = ws[ws.max_row]
            fill = green if abs(diff) < 0.01 else red
            for cell in row:
                cell.fill = fill
        ws.append(["TOTAL", round(pdf_total, 2), round(db_total, 2), round(pdf_total - db_total, 2)])
        autosize(ws)

        summary.append(
            [
                pdf.name,
                week_start,
                week_end,
                round(pdf_total, 2),
                round(db_total, 2),
                round(pdf_total - db_total, 2),
            ]
        )

    autosize(summary)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
