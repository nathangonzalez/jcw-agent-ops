#!/usr/bin/env python3
"""
Payroll reconciliation helper.
Scans exported timesheet XLSX files and compares total hours to SQLite time_entries.
"""

import argparse
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--payroll-root", default=None, help="Path to jcw_payroll repo")
    parser.add_argument("--exports-dir", default=None, help="Path to exports directory")
    parser.add_argument("--db", default=None, help="Path to payroll SQLite db (app.db)")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of files to scan")
    return parser.parse_args()


def find_date(path: Path) -> str | None:
    match = DATE_RE.search(path.name)
    if match:
        return match.group(1)
    match = DATE_RE.search(str(path.parent))
    if match:
        return match.group(1)
    return None


def sum_hours_xlsx(path: Path) -> float:
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = None
    hours_col = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for idx, cell in enumerate(row, start=1):
            if isinstance(cell.value, str) and "hour" in cell.value.lower():
                header_row = cell.row
                hours_col = idx
                break
        if hours_col:
            break
    total = 0.0
    if hours_col:
        for row in ws.iter_rows(min_row=header_row + 1, min_col=hours_col, max_col=hours_col):
            cell = row[0]
            if isinstance(cell.value, (int, float)):
                total += float(cell.value)
        return total

    # Fallback: sum numeric cells in the sheet (best-effort)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                total += float(cell.value)
    return total


def db_hours(conn, start_date: str, end_date: str) -> float:
    cur = conn.cursor()
    cur.execute(
        "SELECT COALESCE(SUM(hours),0) FROM time_entries WHERE work_date >= ? AND work_date <= ?",
        (start_date, end_date),
    )
    return float(cur.fetchone()[0] or 0)


def main():
    args = parse_args()
    payroll_root = Path(args.payroll_root or r"C:\\Users\\natha\\dev\\repos\\jcw_payroll")
    exports_dir = Path(args.exports_dir or payroll_root / "exports")
    db_path = Path(args.db or payroll_root / "data" / "app.db")

    if not exports_dir.exists():
        raise SystemExit(f"Exports directory not found: {exports_dir}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")
    if load_workbook is None:
        raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

    files = sorted(exports_dir.rglob("*.xlsx"))
    if args.limit:
        files = files[: args.limit]

    conn = sqlite3.connect(db_path)
    results = []
    for path in files:
        date_str = find_date(path)
        if not date_str:
            continue
        try:
            week_end = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue
        week_start = week_end - timedelta(days=6)
        export_hours = sum_hours_xlsx(path)
        db_total = db_hours(conn, week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d"))
        results.append(
            {
                "file": str(path),
                "week_end": week_end.strftime("%Y-%m-%d"),
                "export_hours": round(export_hours, 2),
                "db_hours": round(db_total, 2),
                "delta": round(export_hours - db_total, 2),
            }
        )

    conn.close()

    if not results:
        print("No export files with dates found.")
        return

    print("Payroll Reconcile Summary")
    for item in results:
        print(
            f"- {item['week_end']} | export={item['export_hours']} db={item['db_hours']} delta={item['delta']} | {item['file']}"
        )


if __name__ == "__main__":
    main()