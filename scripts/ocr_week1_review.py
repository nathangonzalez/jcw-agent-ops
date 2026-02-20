#!/usr/bin/env python3
"""
Build an OCR review workbook for Week 1 photos.
"""

import argparse
import csv
import sqlite3
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ocr-csv", required=True, help="Path to ocr_combined.csv")
    parser.add_argument("--out", required=True, help="Output XLSX path")
    parser.add_argument("--db", default="", help="Optional payroll DB for employee lookup")
    return parser.parse_args()


def load_employees(db_path: str) -> List[str]:
    if not db_path:
        return []
    path = Path(db_path)
    if not path.exists():
        return []
    conn = sqlite3.connect(str(path))
    cur = conn.cursor()
    cur.execute("SELECT name FROM employees")
    rows = [r[0] for r in cur.fetchall() if r[0]]
    conn.close()
    return rows


def guess_employee(text: str, employees: List[str]) -> str:
    lower = text.lower()
    for name in employees:
        if name.lower() in lower:
            return name
    return ""


def main():
    args = parse_args()
    ocr_csv = Path(args.ocr_csv)
    out_path = Path(args.out)
    if not ocr_csv.exists():
        raise SystemExit(f"Missing OCR CSV: {ocr_csv}")

    employees = load_employees(args.db)

    rows = []
    with ocr_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    wb = Workbook()
    wb.remove(wb.active)

    ocr_sheet = wb.create_sheet("OCR")
    ocr_sheet.append(["image", "avg_confidence", "employee_guess", "text"])
    for r in rows[1:]:
        image, conf, text = r[0], r[1], r[2] if len(r) > 2 else ""
        emp_guess = guess_employee(text, employees)
        ocr_sheet.append([image, conf, emp_guess, text])

    for cell in ocr_sheet[1]:
        cell.font = Font(bold=True)
    ocr_sheet.freeze_panes = "A2"
    ocr_sheet.column_dimensions["A"].width = 22
    ocr_sheet.column_dimensions["B"].width = 14
    ocr_sheet.column_dimensions["C"].width = 20
    ocr_sheet.column_dimensions["D"].width = 120
    for row in ocr_sheet.iter_rows(min_row=2, max_col=4):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")

    review_sheet = wb.create_sheet("Review")
    review_sheet.append(["date", "employee", "customer", "hours", "source_image", "notes", "status"])
    for r in rows[1:]:
        image, conf, text = r[0], r[1], r[2] if len(r) > 2 else ""
        emp_guess = guess_employee(text, employees)
        review_sheet.append(["", emp_guess, "", "", image, "", "needs_review"])

    for cell in review_sheet[1]:
        cell.font = Font(bold=True)
    review_sheet.freeze_panes = "A2"
    for col in range(1, 8):
        review_sheet.column_dimensions[chr(64 + col)].width = 20

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
